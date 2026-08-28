"""Excel exception export, selection, security, and persistence coverage."""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from html import unescape

from openpyxl import load_workbook
from sqlalchemy import event
from werkzeug.datastructures import MultiDict

from app import create_app
from app.extensions import db
from app.models import UsageTotals
from app.services.csv_import import EXPECTED_COLUMNS
from app.services.excel_export import (
    build_exception_workbook,
    load_export_snapshot,
    prepare_export,
    select_export_rows,
)
from app.services.validation_engine import (
    AuditException,
    AuditResult,
    RuleDetail,
    UnableToEvaluate,
)


_EXPECTED_FIXED_HEADERS = (
    "CSV source row",
    "Rule ID",
    "Rule name",
    "Exception message",
    "Active settings profile",
    "RecordID",
    "ApplicationNumber",
    "Gateway",
    "AircraftType",
    "TailNumber",
    "ApplicationDate",
    "StartTime",
    "DateCreated",
    "TruckNumber",
    "Operator",
    "Driver",
)


def _synthetic_export_csv(
    *rows: dict[str, str],
) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(
        output,
        fieldnames=EXPECTED_COLUMNS,
        lineterminator="\n",
    )
    writer.writeheader()

    for index, overrides in enumerate(rows):
        row = {column: "" for column in EXPECTED_COLUMNS}
        row.update(
            {
                "RecordID": f"export-record-{index}",
                "ApplicationNumber": f"export-application-{index}",
                "GatewayCode": "EXPORT-GATEWAY",
                "ApplicationDate": "2026-01-15",
                "StartTime": "08:00",
                "EndTime": "08:30",
                "DateCreated": "2026-01-15 08:00",
                "AircraftType": "2",
                "TailNumber": "AB-123",
                "TruckNumber": "1",
                "Operator": "Export Operator",
                "Driver": "Export Driver",
                "AmbientTemp": "1",
                "Type1Used": "",
                "Type4Used": "",
                "Notes": "Type I applied by truck 2",
            }
        )
        row.update(overrides)
        writer.writerow(row)

    return output.getvalue().encode("utf-8")


def _upload_for_export(client, *rows: dict[str, str]):
    return client.post(
        "/import",
        data={
            "csv_file": (
                io.BytesIO(_synthetic_export_csv(*rows)),
                "export-source.csv",
            )
        },
        content_type="multipart/form-data",
    )


def _export_form(response) -> tuple[str, tuple[str, ...]]:
    html = response.get_data(as_text=True)
    token_match = re.search(
        r'name="export_token" value="([^"]+)"',
        html,
    )
    assert token_match is not None
    identifiers = tuple(
        re.findall(r'name="exception_id"\s+value="([^"]+)"', html)
    )
    return unescape(token_match.group(1)), identifiers


def _workbook_from_response(response):
    return load_workbook(io.BytesIO(response.data))


def _header_positions(worksheet) -> dict[str, int]:
    return {
        cell.value: index
        for index, cell in enumerate(worksheet[1], start=1)
    }


def _audit_exception(
    *,
    source_row_number: int = 2,
    rule_id: str = "CC-RULE-001",
    rule_name: str = "Application Entry Proceeds Event",
    exception_message: str = "Application entry proceeds event.",
    record_id: str = "record-001",
    application_number: str = "application-001",
    gateway_code: str = "GATEWAY-A",
    aircraft_type: str = "2",
    tail_number: str = "AB-123",
    application_date: str = "2026-01-15",
    start_time: str = "08:00",
    date_created: str = "2026-01-15 07:59",
    truck_number: str = "1",
    operator: str = "Operator",
    driver: str = "Driver",
    details: tuple[RuleDetail, ...] = (
        RuleDetail("Timing difference", "1 minute"),
    ),
) -> AuditException:
    return AuditException(
        rule_id=rule_id,
        rule_name=rule_name,
        exception_message=exception_message,
        source_row_number=source_row_number,
        record_id=record_id,
        application_number=application_number,
        gateway_code=gateway_code,
        aircraft_type=aircraft_type,
        tail_number=tail_number,
        application_date=application_date,
        start_time=start_time,
        date_created=date_created,
        truck_number=truck_number,
        operator=operator,
        driver=driver,
        details=details,
    )


def _audit_result(
    *exceptions: AuditException,
    warnings: tuple[UnableToEvaluate, ...] = (),
) -> AuditResult:
    return AuditResult(
        filename="request-only.csv",
        rows_audited=2,
        rules_executed=15,
        active_settings_profile_name="Default",
        exceptions=exceptions,
        unable_to_evaluate=warnings,
    )


def test_results_show_one_checkbox_per_exception_and_export_controls(client):
    response = _upload_for_export(
        client,
        {
            "DateCreated": "2026-01-15 07:59",
            "TailNumber": "N121UP",
        },
    )

    assert response.status_code == 200
    assert response.data.count(b'data-exception-checkbox') == 2
    assert b"Select All" in response.data
    assert b"Clear All" in response.data
    assert b"Export Selected" in response.data
    assert b"Export All" in response.data
    assert b"Export Exceptions" not in response.data
    assert b'href="#audit-results"' in response.data
    assert b'id="audit-results"' in response.data
    desktop_nav = re.search(
        rb'<nav class="site-nav".*?</nav>',
        response.data,
        flags=re.DOTALL,
    )
    assert desktop_nav is not None
    assert re.findall(
        rb">\s*(Import|Rules|Reports|Settings|Sign In|Create Account)\s*<",
        desktop_nav.group(0),
    ) == [
        b"Import",
        b"Rules",
        b"Reports",
        b"Settings",
        b"Sign In",
        b"Create Account",
    ]
    assert re.search(
        rb'class="site-nav__link site-nav__link--active"'
        rb'[^>]+href="#audit-results"[^>]+aria-current="page"',
        response.data,
    )
    assert re.search(
        rb'<button[^>]+name="scope"[^>]+value="all"[^>]+data-export-all',
        response.data,
    )
    assert re.search(
        rb'value="selected"\s+disabled\s+data-export-selected',
        response.data,
    )
    token, identifiers = _export_form(response)
    assert token
    assert identifiers == ("exception-1", "exception-2")


def test_export_navigation_is_hidden_when_audit_has_no_exceptions(client):
    response = _upload_for_export(client, {})

    assert response.status_code == 200
    assert b"No exceptions found" in response.data
    assert b'href="#audit-results"' in response.data
    assert re.search(rb">\s*Reports\s*</a>", response.data)
    assert b"Export Exceptions" not in response.data
    assert b"data-export-all" not in response.data


def test_export_all_downloads_every_exception_in_audit_order(client):
    results = _upload_for_export(
        client,
        {
            "DateCreated": "2026-01-15 07:59",
            "TailNumber": "N121UP",
        },
        {
            "DateCreated": "2026-01-15 07:58",
        },
    )
    token, identifiers = _export_form(results)

    response = client.post(
        "/export",
        data=MultiDict(
            (
                ("export_token", token),
                ("scope", "all"),
                *(("exception_id", identifier) for identifier in identifiers),
            )
        ),
    )

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert re.search(
        r'attachment; filename=CryoCheck_Exceptions_\d{8}_\d{6}\.xlsx',
        response.headers["Content-Disposition"],
    )
    assert response.headers["Cache-Control"] == "no-store"
    workbook = _workbook_from_response(response)
    worksheet = workbook["Exceptions"]
    headers = _header_positions(worksheet)
    exported_order = tuple(
        (
            worksheet.cell(row=row, column=headers["CSV source row"]).value,
            worksheet.cell(row=row, column=headers["Rule ID"]).value,
        )
        for row in range(2, worksheet.max_row + 1)
    )

    assert exported_order == (
        (2, "CC-RULE-001"),
        (2, "CC-RULE-012"),
        (3, "CC-RULE-001"),
    )
    workbook.close()


def test_export_selected_ignores_submission_order_and_keeps_audit_order(client):
    results = _upload_for_export(
        client,
        {
            "DateCreated": "2026-01-15 07:59",
            "TailNumber": "N121UP",
        },
        {
            "DateCreated": "2026-01-15 07:58",
        },
    )
    token, identifiers = _export_form(results)

    response = client.post(
        "/export",
        data=MultiDict(
            (
                ("export_token", token),
                ("scope", "selected"),
                ("exception_id", identifiers[2]),
                ("exception_id", identifiers[0]),
            )
        ),
    )
    workbook = _workbook_from_response(response)
    worksheet = workbook["Exceptions"]
    headers = _header_positions(worksheet)

    assert response.status_code == 200
    assert tuple(
        (
            worksheet.cell(row=row, column=headers["CSV source row"]).value,
            worksheet.cell(row=row, column=headers["Rule ID"]).value,
        )
        for row in range(2, worksheet.max_row + 1)
    ) == ((2, "CC-RULE-001"), (3, "CC-RULE-001"))
    workbook.close()


def test_ios_validation_accepts_selected_and_all_without_building_workbook(
    client,
    monkeypatch,
):
    results = _upload_for_export(
        client,
        {
            "DateCreated": "2026-01-15 07:59",
            "TailNumber": "N121UP",
        },
        {
            "DateCreated": "2026-01-15 07:58",
        },
    )
    token, identifiers = _export_form(results)

    def fail_if_workbook_is_built(_selected_rows):
        raise AssertionError("Validation must not build or persist a workbook")

    monkeypatch.setattr(
        "app.routes.build_exception_workbook",
        fail_if_workbook_is_built,
    )
    selected = client.post(
        "/export",
        data=MultiDict(
            (
                ("export_token", token),
                ("scope", "selected"),
                ("exception_id", identifiers[2]),
                ("exception_id", identifiers[0]),
                ("delivery", "validate"),
            )
        ),
    )
    all_rows = client.post(
        "/export",
        data={
            "export_token": token,
            "scope": "all",
            "delivery": "validate",
        },
    )

    assert selected.status_code == 200
    assert selected.get_json() == {"ok": True, "selected_count": 2}
    assert selected.headers["Cache-Control"] == "no-store"
    assert all_rows.status_code == 200
    assert all_rows.get_json() == {"ok": True, "selected_count": 3}
    assert all_rows.headers["Cache-Control"] == "no-store"


def test_ios_native_delivery_streams_the_existing_complete_workbook(client):
    results = _upload_for_export(
        client,
        {
            "DateCreated": "2026-01-15 07:59",
            "TailNumber": "N121UP",
        },
    )
    token, identifiers = _export_form(results)

    response = client.post(
        "/export",
        data={
            "export_token": token,
            "scope": "selected",
            "exception_id": identifiers[0],
            "delivery": "native",
        },
    )

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["Cache-Control"] == "no-store"
    workbook = _workbook_from_response(response)
    worksheet = workbook["Exceptions"]
    headers = _header_positions(worksheet)
    assert worksheet.cell(row=2, column=headers["CSV source row"]).value == 2
    assert worksheet.cell(row=2, column=headers["Rule ID"]).value == (
        "CC-RULE-001"
    )
    workbook.close()


def test_export_selected_requires_at_least_one_selection(client):
    results = _upload_for_export(
        client,
        {"DateCreated": "2026-01-15 07:59"},
    )
    token, _identifiers = _export_form(results)

    response = client.post(
        "/export",
        data={"export_token": token, "scope": "selected"},
    )

    assert response.status_code == 400
    assert b"The exception export could not be created" in response.data
    assert b"Select at least one exception" in response.data


def test_ios_validation_returns_specific_json_for_selection_and_token_errors(
    app,
    client,
):
    results = _upload_for_export(
        client,
        {"DateCreated": "2026-01-15 07:59"},
    )
    token, identifiers = _export_form(results)

    missing_selection = client.post(
        "/export",
        data={
            "export_token": token,
            "scope": "selected",
            "delivery": "validate",
        },
    )
    malformed = client.post(
        "/export",
        data={
            "export_token": f"{token}tampered",
            "scope": "all",
            "delivery": "validate",
        },
    )
    app.config["EXPORT_TOKEN_MAX_AGE_SECONDS"] = -1
    expired = client.post(
        "/export",
        data={
            "export_token": token,
            "scope": "selected",
            "exception_id": identifiers[0],
            "delivery": "validate",
        },
    )

    assert missing_selection.status_code == 400
    assert missing_selection.get_json() == {
        "ok": False,
        "error": "Select at least one exception before exporting.",
    }
    assert malformed.status_code == 400
    assert "export request is invalid" in malformed.get_json()["error"]
    assert expired.status_code == 400
    assert "export request expired" in expired.get_json()["error"]
    for response in (missing_selection, malformed, expired):
        assert response.headers["Cache-Control"] == "no-store"


def test_export_rejects_unknown_and_duplicate_identifiers(client):
    results = _upload_for_export(
        client,
        {"DateCreated": "2026-01-15 07:59"},
    )
    token, identifiers = _export_form(results)

    unknown = client.post(
        "/export",
        data={
            "export_token": token,
            "scope": "selected",
            "exception_id": "exception-999",
        },
    )
    duplicate = client.post(
        "/export",
        data=MultiDict(
            (
                ("export_token", token),
                ("scope", "selected"),
                ("exception_id", identifiers[0]),
                ("exception_id", identifiers[0]),
            )
        ),
    )

    assert unknown.status_code == 400
    assert b"not part of this audit result" in unknown.data
    assert duplicate.status_code == 400
    assert b"duplicate exception selections" in duplicate.data


def test_export_rejects_malformed_and_expired_snapshots(app, client):
    results = _upload_for_export(
        client,
        {"DateCreated": "2026-01-15 07:59"},
    )
    token, identifiers = _export_form(results)

    malformed = client.post(
        "/export",
        data={
            "export_token": f"{token}tampered",
            "scope": "all",
        },
    )
    app.config["EXPORT_TOKEN_MAX_AGE_SECONDS"] = -1
    expired = client.post(
        "/export",
        data={
            "export_token": token,
            "scope": "selected",
            "exception_id": identifiers[0],
        },
    )

    assert malformed.status_code == 400
    assert b"export request is invalid" in malformed.data
    assert expired.status_code == 400
    assert b"export request expired" in expired.data


def test_new_import_invalidates_an_older_results_export(client):
    older_results = _upload_for_export(
        client,
        {"DateCreated": "2026-01-15 07:59"},
    )
    older_token, _older_identifiers = _export_form(older_results)
    newer_results = _upload_for_export(
        client,
        {"TailNumber": "N121UP"},
    )
    newer_token, _newer_identifiers = _export_form(newer_results)

    stale = client.post(
        "/export",
        data={"export_token": older_token, "scope": "all"},
    )
    current = client.post(
        "/export",
        data={"export_token": newer_token, "scope": "all"},
    )

    assert stale.status_code == 400
    assert b"no longer the current audit result" in stale.data
    assert current.status_code == 200


def test_export_workbook_has_required_headers_values_and_formatting():
    audit = _audit_result(
        _audit_exception(
            details=(
                RuleDetail("Timing difference", "1 minute"),
                RuleDetail("Comparison", "Entry preceded event."),
            )
        )
    )
    prepared = prepare_export(
        audit,
        secret_key="export-test-secret",
        context_id="export-test-context",
    )
    snapshot = load_export_snapshot(
        prepared.token,
        secret_key="export-test-secret",
        max_age_seconds=60,
        expected_context_id="export-test-context",
    )
    selected = select_export_rows(
        snapshot,
        scope="all",
        selected_identifiers=(),
    )
    stream, filename = build_exception_workbook(
        selected,
        now=datetime(2026, 7, 23, 12, 34, 56, tzinfo=timezone.utc),
    )
    workbook = load_workbook(stream)
    worksheet = workbook["Exceptions"]
    headers = tuple(cell.value for cell in worksheet[1])
    positions = _header_positions(worksheet)

    assert filename == "CryoCheck_Exceptions_20260723_123456.xlsx"
    assert headers[: len(_EXPECTED_FIXED_HEADERS)] == _EXPECTED_FIXED_HEADERS
    assert "Detail — Timing difference" in headers
    assert "Detail — Comparison" in headers
    assert headers[-1] == "Combined details"
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == worksheet.dimensions
    assert all(cell.font.bold for cell in worksheet[1])
    assert all(cell.alignment.wrap_text for cell in worksheet[1])
    assert worksheet["A2"].value == 2
    assert worksheet.cell(
        row=2,
        column=positions["Active settings profile"],
    ).value == "Default"
    assert worksheet.cell(
        row=2,
        column=positions["Detail — Timing difference"],
    ).value == "1 minute"
    assert worksheet.cell(
        row=2,
        column=positions["Combined details"],
    ).value == (
        "Timing difference: 1 minute; Comparison: Entry preceded event."
    )
    assert worksheet.column_dimensions["A"].width >= 12
    workbook.close()


def test_simplified_result_fields_remain_complete_in_excel():
    audit = _audit_result(
        _audit_exception(
            source_row_number=2,
            rule_id="CC-RULE-005",
            rule_name="BRIX Out of Range",
            details=(
                RuleDetail("Entered BRIX", "50"),
                RuleDetail("Selected Type IV fluid", "Synthetic Fluid"),
                RuleDetail("Acceptable inclusive range", "34.6\N{EN DASH}36.6"),
                RuleDetail("Range comparison", "Above range"),
                RuleDetail("Amount above nearest boundary", "13.4"),
            ),
        ),
        _audit_exception(
            source_row_number=3,
            rule_id="CC-RULE-006",
            rule_name="Excessive Gap Between Steps",
            details=(
                RuleDetail("Type I end time", "08:00"),
                RuleDetail("Type IV start time", "08:06"),
                RuleDetail("Actual calculated gap", "6 minutes"),
                RuleDetail("Configured Allowed Gap", "5 minutes"),
                RuleDetail("Amount over setting", "1 minute"),
            ),
        ),
        _audit_exception(
            source_row_number=4,
            rule_id="CC-RULE-010",
            rule_name="Excessive Event Time",
            details=(
                RuleDetail("ProcessTime1", "20 minutes"),
                RuleDetail("ProcessTime4", "11 minutes"),
                RuleDetail("Include Gap setting", "Off"),
                RuleDetail("Calculated event time", "31 minutes"),
                RuleDetail("Configured maximum event time", "30 minutes"),
                RuleDetail("Minutes over the maximum", "1 minute"),
            ),
        ),
    )
    prepared = prepare_export(
        audit,
        secret_key="complete-details-secret",
        context_id="complete-details-context",
    )
    snapshot = load_export_snapshot(
        prepared.token,
        secret_key="complete-details-secret",
        max_age_seconds=60,
        expected_context_id="complete-details-context",
    )
    selected = select_export_rows(
        snapshot,
        scope="all",
        selected_identifiers=(),
    )
    stream, _filename = build_exception_workbook(selected)
    workbook = load_workbook(stream)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Range comparison"],
    ).value == "Above range"
    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Amount above nearest boundary"],
    ).value == "13.4"
    assert worksheet.cell(
        row=3,
        column=positions["Detail \N{EM DASH} Amount over setting"],
    ).value == "1 minute"
    assert worksheet.cell(
        row=4,
        column=positions["Detail \N{EM DASH} Minutes over the maximum"],
    ).value == "1 minute"
    assert "Minutes over the maximum: 1 minute" in worksheet.cell(
        row=4,
        column=positions["Combined details"],
    ).value
    workbook.close()


def test_rate_and_concentration_details_remain_complete_in_excel():
    audit = _audit_result(
        _audit_exception(
            source_row_number=2,
            rule_id="CC-RULE-008",
            rule_name="Excessive Type I",
            details=(
                RuleDetail("Type I gallons used", "121.00 gallons"),
                RuleDetail("Recorded ProcessTime1", "1.0 minute"),
                RuleDetail("Adjusted calculation time", "2 minutes"),
                RuleDetail(
                    "Adjusted Type I rate",
                    "60.5 gallons per minute",
                ),
                RuleDetail(
                    "Configured maximum Type I rate",
                    "60 gallons per minute",
                ),
                RuleDetail("Comparison", "Rate exceeds maximum."),
            ),
        ),
        _audit_exception(
            source_row_number=3,
            rule_id="CC-RULE-009",
            rule_name="Excessive Type IV",
            details=(
                RuleDetail("Type IV gallons used", "61.00 gallons"),
                RuleDetail("Recorded ProcessTime4", "1.0 minute"),
                RuleDetail("Adjusted calculation time", "2 minutes"),
                RuleDetail(
                    "Adjusted Type IV rate",
                    "30.5 gallons per minute",
                ),
                RuleDetail(
                    "Configured maximum Type IV rate",
                    "30 gallons per minute",
                ),
                RuleDetail("Comparison", "Rate exceeds maximum."),
            ),
        ),
        _audit_exception(
            source_row_number=4,
            rule_id="CC-RULE-011",
            rule_name="Incorrect Type IV Concentration",
            details=(
                RuleDetail("Selected Type IV fluid", "Synthetic Fluid"),
                RuleDetail("Entered Type IV concentration", "99.9"),
                RuleDetail("Required Type IV concentration", "100%"),
                RuleDetail(
                    "Comparison",
                    "Entered concentration must equal 100%.",
                ),
            ),
        ),
    )
    prepared = prepare_export(
        audit,
        secret_key="rate-details-secret",
        context_id="rate-details-context",
    )
    snapshot = load_export_snapshot(
        prepared.token,
        secret_key="rate-details-secret",
        max_age_seconds=60,
        expected_context_id="rate-details-context",
    )
    selected = select_export_rows(
        snapshot,
        scope="all",
        selected_identifiers=(),
    )
    stream, _filename = build_exception_workbook(selected)
    workbook = load_workbook(stream)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Adjusted calculation time"],
    ).value == "2 minutes"
    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Comparison"],
    ).value == "Rate exceeds maximum."
    assert worksheet.cell(
        row=3,
        column=positions["Detail \N{EM DASH} Adjusted calculation time"],
    ).value == "2 minutes"
    assert worksheet.cell(
        row=3,
        column=positions["Detail \N{EM DASH} Comparison"],
    ).value == "Rate exceeds maximum."
    assert worksheet.cell(
        row=4,
        column=positions["Detail \N{EM DASH} Selected Type IV fluid"],
    ).value == "Synthetic Fluid"
    assert worksheet.cell(
        row=4,
        column=positions["Detail \N{EM DASH} Comparison"],
    ).value == "Entered concentration must equal 100%."
    assert "Selected Type IV fluid: Synthetic Fluid" in worksheet.cell(
        row=4,
        column=positions["Combined details"],
    ).value
    workbook.close()


def test_final_simplified_details_remain_complete_in_excel():
    audit = _audit_result(
        _audit_exception(
            source_row_number=2,
            rule_id="CC-RULE-001",
            rule_name="Application Entry Proceeds Event",
            details=(
                RuleDetail("Application date/time", "2026-01-15 08:00"),
                RuleDetail("Entry date/time", "2026-01-15 07:59"),
                RuleDetail(
                    "How far before the application event the entry was created",
                    "1 minute",
                ),
            ),
        ),
        _audit_exception(
            source_row_number=3,
            rule_id="CC-RULE-003",
            rule_name="Incorrect Freeze Point",
            details=(
                RuleDetail("Selected Type I fluid", "Synthetic Type I"),
                RuleDetail("Recorded concentration", "65%"),
                RuleDetail("Entered freeze point", "-20 F"),
                RuleDetail(
                    "Expected manufacturer-chart freeze point",
                    "-50.0 F",
                ),
                RuleDetail(
                    "Comparison",
                    "Expected -50.0 F; entered -20 F.",
                ),
            ),
        ),
        _audit_exception(
            source_row_number=4,
            rule_id="CC-RULE-014",
            rule_name="Type IV Without Type I Explanation Required",
            details=(
                RuleDetail("AircraftType", "2"),
                RuleDetail("Type1Used", ""),
                RuleDetail("Type4Used", "1"),
                RuleDetail("Current TruckNumber", "1"),
                RuleDetail("Original Notes", "Type I applied by truck 1"),
                RuleDetail(
                    "Missing or failed requirement",
                    "Documented truck number matches current TruckNumber",
                ),
                RuleDetail("Documented truck number", "1"),
            ),
        ),
    )
    prepared = prepare_export(
        audit,
        secret_key="final-details-secret",
        context_id="final-details-context",
    )
    snapshot = load_export_snapshot(
        prepared.token,
        secret_key="final-details-secret",
        max_age_seconds=60,
        expected_context_id="final-details-context",
    )
    selected = select_export_rows(
        snapshot,
        scope="all",
        selected_identifiers=(),
    )
    stream, _filename = build_exception_workbook(selected)
    workbook = load_workbook(stream)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Entry date/time"],
    ).value == "2026-01-15 07:59"
    assert worksheet.cell(
        row=3,
        column=positions["Detail \N{EM DASH} Selected Type I fluid"],
    ).value == "Synthetic Type I"
    assert worksheet.cell(
        row=3,
        column=positions["Detail \N{EM DASH} Comparison"],
    ).value == "Expected -50.0 F; entered -20 F."
    assert worksheet.cell(
        row=4,
        column=positions[
            "Detail \N{EM DASH} Missing or failed requirement"
        ],
    ).value == "Documented truck number matches current TruckNumber"
    assert worksheet.cell(
        row=4,
        column=positions["Detail \N{EM DASH} Current TruckNumber"],
    ).value == "1"
    assert "Missing or failed requirement" in worksheet.cell(
        row=4,
        column=positions["Combined details"],
    ).value
    workbook.close()


def test_rule_guidance_details_remain_complete_in_excel():
    audit = _audit_result(
        _audit_exception(
            source_row_number=2,
            rule_id="CC-RULE-007",
            rule_name="No Type IV During Active Precipitation",
            details=(
                RuleDetail("Recorded precipitation", "Snow"),
                RuleDetail("Type IV amount recorded", "0"),
                RuleDetail("Finding", "No positive Type IV usage was recorded."),
            ),
        ),
        _audit_exception(
            source_row_number=3,
            rule_id="CC-RULE-012",
            rule_name="Incorrect Tail Number",
            details=(
                RuleDetail("Original AircraftType", "1"),
                RuleDetail("Original TailNumber", "AB-123"),
                RuleDetail("Required format", "UPS NxxxUP format"),
                RuleDetail(
                    "Failure reason",
                    "Does not match UPS NxxxUP format",
                ),
            ),
        ),
        _audit_exception(
            source_row_number=4,
            rule_id="CC-RULE-013",
            rule_name="Pass Overlap",
            details=(
                RuleDetail("Overall StartTime", "20:00"),
                RuleDetail("Overall EndTime", "20:30"),
                RuleDetail("Type I EndTime1", "20:20"),
                RuleDetail("Type IV StartTime4", "20:15"),
                RuleDetail("Calculated overlap", "5 minutes"),
                RuleDetail(
                    "Explanation",
                    "Type IV began before Type I ended.",
                ),
            ),
        ),
    )
    prepared = prepare_export(
        audit,
        secret_key="guidance-details-secret",
        context_id="guidance-details-context",
    )
    snapshot = load_export_snapshot(
        prepared.token,
        secret_key="guidance-details-secret",
        max_age_seconds=60,
        expected_context_id="guidance-details-context",
    )
    selected = select_export_rows(
        snapshot,
        scope="all",
        selected_identifiers=(),
    )
    stream, _filename = build_exception_workbook(selected)
    workbook = load_workbook(stream)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Finding"],
    ).value == "No positive Type IV usage was recorded."
    assert worksheet.cell(
        row=3,
        column=positions["Detail \N{EM DASH} Required format"],
    ).value == "UPS NxxxUP format"
    assert worksheet.cell(
        row=4,
        column=positions["Detail \N{EM DASH} Calculated overlap"],
    ).value == "5 minutes"
    assert worksheet.cell(
        row=4,
        column=positions["Detail \N{EM DASH} Explanation"],
    ).value == "Type IV began before Type I ended."
    assert "Calculated overlap: 5 minutes" in worksheet.cell(
        row=4,
        column=positions["Combined details"],
    ).value
    workbook.close()


def test_results_and_export_preserve_non_padded_start_time_text(client):
    results = _upload_for_export(
        client,
        {
            "ApplicationDate": "1/1/2026",
            "StartTime": "5:11",
            "DateCreated": "1/2/2026 8:08",
        },
    )

    assert results.status_code == 200
    assert b"Application Date" in results.data
    assert b"1/1/2026" in results.data
    assert b"Unable to evaluate" not in results.data
    token, _identifiers = _export_form(results)
    response = client.post(
        "/export",
        data={"export_token": token, "scope": "all"},
    )
    workbook = _workbook_from_response(response)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert response.status_code == 200
    assert worksheet.cell(row=2, column=positions["Rule ID"]).value == (
        "CC-RULE-002"
    )
    assert worksheet.cell(row=2, column=positions["StartTime"]).value == (
        "5:11"
    )
    workbook.close()


def test_outside_chart_concentration_details_are_exported(client):
    results = _upload_for_export(
        client,
        {
            "RecordID": "422634",
            "Type1Used": "10",
            "Type1Concentration": "90",
        },
    )
    token, identifiers = _export_form(results)

    assert identifiers == ("exception-1",)
    response = client.post(
        "/export",
        data={"export_token": token, "scope": "all"},
    )
    workbook = _workbook_from_response(response)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert response.status_code == 200
    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=positions["Rule ID"]).value == (
        "CC-RULE-003"
    )
    assert worksheet.cell(
        row=2,
        column=positions["Exception message"],
    ).value == "Type I concentration outside manufacturer chart."
    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Entered concentration"],
    ).value == "90%"
    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Selected Type I fluid"],
    ).value == "Cryotech Polar Plus LT"
    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Supported chart range"],
    ).value == "0\N{EN DASH}70%"
    assert worksheet.cell(
        row=2,
        column=positions["Detail \N{EM DASH} Comparison"],
    ).value == (
        "Entered Type I concentration 90% is outside the supported "
        "manufacturer-chart range of 0\N{EN DASH}70%."
    )
    assert "Entered concentration: 90%" in worksheet.cell(
        row=2,
        column=positions["Combined details"],
    ).value
    workbook.close()


def test_export_escapes_formula_like_text_in_source_and_detail_fields():
    audit = _audit_result(
        _audit_exception(
            record_id="=SUM(A1:A2)",
            application_number="+1",
            gateway_code="-2",
            tail_number="@command",
            details=(
                RuleDetail("Entered value", "=HYPERLINK(\"bad\")"),
                RuleDetail("Negative text", "-1"),
            ),
        )
    )
    prepared = prepare_export(
        audit,
        secret_key="formula-test-secret",
        context_id="formula-test-context",
    )
    snapshot = load_export_snapshot(
        prepared.token,
        secret_key="formula-test-secret",
        max_age_seconds=60,
        expected_context_id="formula-test-context",
    )
    stream, _filename = build_exception_workbook(
        snapshot.rows,
        now=datetime(2026, 7, 23, tzinfo=timezone.utc),
    )
    workbook = load_workbook(stream, data_only=False)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert worksheet.cell(row=2, column=positions["RecordID"]).value == (
        "'=SUM(A1:A2)"
    )
    assert worksheet.cell(
        row=2,
        column=positions["ApplicationNumber"],
    ).value == "'+1"
    assert worksheet.cell(row=2, column=positions["Gateway"]).value == "'-2"
    assert worksheet.cell(row=2, column=positions["TailNumber"]).value == (
        "'@command"
    )
    assert worksheet.cell(
        row=2,
        column=positions["Detail — Entered value"],
    ).value == "'=HYPERLINK(\"bad\")"
    assert worksheet.cell(
        row=2,
        column=positions["Detail — Negative text"],
    ).value == "'-1"
    assert all(
        cell.data_type != "f"
        for row in worksheet.iter_rows(min_row=2)
        for cell in row
    )
    workbook.close()


def test_unable_to_evaluate_warnings_are_excluded_from_export(client):
    results = _upload_for_export(
        client,
        {
            "DateCreated": "malformed",
            "TailNumber": "N121UP",
        },
    )

    assert b"Some rule evaluations could not run" in results.data
    token, _identifiers = _export_form(results)
    response = client.post(
        "/export",
        data={"export_token": token, "scope": "all"},
    )
    workbook = _workbook_from_response(response)
    worksheet = workbook["Exceptions"]
    positions = _header_positions(worksheet)

    assert response.status_code == 200
    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=positions["Rule ID"]).value == (
        "CC-RULE-012"
    )
    assert "unable" not in " ".join(
        str(cell.value or "")
        for row in worksheet.iter_rows()
        for cell in row
    ).lower()
    workbook.close()


def test_anonymous_export_writes_only_its_aggregate_usage_total(app, client):
    results = _upload_for_export(
        client,
        {"DateCreated": "2026-01-15 07:59"},
    )
    token, _identifiers = _export_form(results)
    executed_statements: list[str] = []

    def record_statement(
        connection,
        cursor,
        statement,
        parameters,
        context,
        executemany,
    ):
        del connection, cursor, parameters, context, executemany
        executed_statements.append(statement)

    with app.app_context():
        engine = db.engine
        event.listen(engine, "before_cursor_execute", record_statement)
        try:
            response = client.post(
                "/export",
                data={"export_token": token, "scope": "all"},
            )
        finally:
            event.remove(engine, "before_cursor_execute", record_statement)

    assert response.status_code == 200
    mutating = [
        statement
        for statement in executed_statements
        if statement.lstrip().upper().startswith(
            ("INSERT", "UPDATE", "DELETE", "CREATE", "DROP", "ALTER")
        )
    ]
    assert len(mutating) == 1
    assert "usage_totals" in mutating[0]
    with app.app_context():
        totals = db.session.get(UsageTotals, 1)
        assert totals is not None
        assert totals.anonymous_export_count == 1


def test_production_export_route_requires_csrf():
    production_app = create_app("production")
    client = production_app.test_client()

    response = client.post(
        "/export",
        base_url="https://localhost",
        data={"export_token": "not-reached", "scope": "all"},
    )

    assert response.status_code == 400
    assert b"Security check failed" in response.data


def test_ios_validation_preserves_csrf_cookie_and_signed_session_context(
    app,
    client,
):
    app.config["WTF_CSRF_ENABLED"] = True
    landing = client.get("/", base_url="https://localhost")
    upload_csrf = re.search(
        rb'name="csrf_token" value="([^"]+)"',
        landing.data,
    )
    assert upload_csrf is not None

    results = client.post(
        "/import",
        base_url="https://localhost",
        headers={"Referer": "https://localhost/"},
        data={
            "csrf_token": upload_csrf.group(1).decode(),
            "csv_file": (
                io.BytesIO(
                    _synthetic_export_csv(
                        {"DateCreated": "2026-01-15 07:59"}
                    )
                ),
                "ios-export.csv",
            ),
        },
        content_type="multipart/form-data",
    )
    export_html = results.get_data(as_text=True)
    export_csrf = re.search(
        r'id="exception-export-form".*?'
        r'name="csrf_token" value="([^"]+)"',
        export_html,
        flags=re.DOTALL,
    )
    export_token, identifiers = _export_form(results)

    assert results.status_code == 200
    assert export_csrf is not None
    validation = client.post(
        "/export",
        base_url="https://localhost",
        headers={"Referer": "https://localhost/import"},
        data={
            "csrf_token": export_csrf.group(1),
            "export_token": export_token,
            "scope": "selected",
            "exception_id": identifiers[0],
            "delivery": "validate",
        },
    )

    assert validation.status_code == 200
    assert validation.get_json() == {"ok": True, "selected_count": 1}
    assert validation.headers["Cache-Control"] == "no-store"
