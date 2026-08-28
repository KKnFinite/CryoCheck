"""Focused coverage for CC-RULE-015 minimum spray-amount validation."""

from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.services.csv_import import CSVImportResult, CSVSourceRow
from app.services.excel_export import (
    build_exception_workbook,
    load_export_snapshot,
    prepare_export,
    select_export_rows,
)
from app.services.results_display import exception_presentation
from app.services.settings import DEFAULT_SETTINGS
from app.services.validation_engine import run_audit


def _audit(*, settings=DEFAULT_SETTINGS, **overrides):
    fields = {
        "RecordID": "minimum-rate-record",
        "ApplicationNumber": "minimum-rate-application",
        "GatewayCode": "TEST",
        "ApplicationDate": "2026-08-27",
        "StartTime": "08:00",
        "EndTime": "08:30",
        "DateCreated": "2026-08-27 08:00",
        "AircraftType": "2",
        "TailNumber": "AB-123",
        "TruckNumber": "1",
        "Type1Used": "",
        "ProcessTime1": "1",
        "Type4Used": "",
        "ProcessTime4": "1",
    }
    fields.update(overrides)
    row = CSVSourceRow(source_row_number=2, fields=tuple(fields.items()))
    imported = CSVImportResult(
        filename="minimum-rate.csv",
        row_count=1,
        column_count=len(fields),
        column_names=tuple(fields),
        rows=(row,),
        expected_columns_found=(),
        missing_columns=(),
        unexpected_columns=(),
        gateway_codes=("TEST",),
        earliest_application_date="2026-08-27",
        latest_application_date="2026-08-27",
    )
    return run_audit(imported, settings)


def _rule_015_exceptions(audit):
    return tuple(item for item in audit.exceptions if item.rule_id == "CC-RULE-015")


def _rule_015_warnings(audit):
    return tuple(item for item in audit.unable_to_evaluate if item.rule_id == "CC-RULE-015")


def test_default_minimum_gallon_boundaries_for_type_i_and_type_iv():
    assert len(_rule_015_exceptions(_audit(Type1Used="0.99"))) == 1
    assert not _rule_015_exceptions(_audit(Type1Used="1"))
    assert not _rule_015_exceptions(_audit(Type1Used="1.01"))

    assert len(_rule_015_exceptions(_audit(Type4Used="4.99"))) == 1
    assert not _rule_015_exceptions(_audit(Type4Used="5"))
    assert not _rule_015_exceptions(_audit(Type4Used="5.01"))


def test_personal_minimum_gallon_overrides_apply_independently():
    settings = replace(
        DEFAULT_SETTINGS,
        name="Personal — MinimumRateUser",
        is_default=False,
        min_type1_gallons=Decimal("2"),
        min_type4_gallons=Decimal("8"),
    )

    assert not _rule_015_exceptions(
        _audit(settings=settings, Type1Used="2")
    )
    assert not _rule_015_exceptions(
        _audit(settings=settings, Type4Used="8")
    )


def test_migration_uses_new_defaults_without_copying_old_rate_values():
    migration = Path(
        "migrations/versions/f2a6b8c1d904_correct_minimum_spray_amount_settings.py"
    ).read_text(encoding="utf-8")

    assert '"min_type1_gallons"' in migration
    assert 'server_default="1"' in migration
    assert '"min_type4_gallons"' in migration
    assert 'server_default="5"' in migration
    assert "UPDATE user_settings" not in migration


def test_both_fluids_can_fail_independently_on_one_row():
    audit = _audit(Type1Used="0.5", Type4Used="4")

    exceptions = _rule_015_exceptions(audit)
    assert len(exceptions) == 2
    assert [exception.details[0].label for exception in exceptions] == [
        "Type I gallons used",
        "Type IV gallons used",
    ]


def test_malformed_values_warn_and_blank_or_nonpositive_usage_skips():
    malformed_usage = _audit(Type1Used="NaN")
    assert _rule_015_warnings(malformed_usage)[0].invalid_fields == ("Type1Used",)

    skipped = _audit(Type1Used="0", ProcessTime1="bad", Type4Used="", ProcessTime4="bad")
    assert not _rule_015_exceptions(skipped)
    assert not _rule_015_warnings(skipped)


def test_rule_015_no_longer_depends_on_process_time():
    audit = _audit(
        Type1Used="0.5",
        ProcessTime1="bad",
        Type4Used="4",
        ProcessTime4="NaN",
    )
    assert len(_rule_015_exceptions(audit)) == 2
    assert not _rule_015_warnings(audit)


def test_results_card_is_concise_and_marks_failing_usage_invalid():
    exception = _rule_015_exceptions(_audit(Type1Used="0.5"))[0]

    presentation = exception_presentation(exception)
    assert tuple(detail.label for detail in presentation.details) == (
        "Type I Used",
        "Minimum Type I Gallons",
    )
    assert presentation.details[0].value_kind == "invalid"


def test_excel_export_retains_complete_minimum_amount_details():
    audit = _audit(Type1Used="0.5")
    prepared = prepare_export(audit, secret_key="minimum-rate-secret", context_id="minimum-rate")
    snapshot = load_export_snapshot(
        prepared.token,
        secret_key="minimum-rate-secret",
        max_age_seconds=60,
        expected_context_id="minimum-rate",
    )
    stream, _filename = build_exception_workbook(
        select_export_rows(snapshot, scope="all", selected_identifiers=())
    )
    workbook = load_workbook(stream)
    worksheet = workbook["Exceptions"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    assert worksheet.cell(2, headers["Detail — Configured minimum Type I gallons"]).value == "1 gallon"
    assert "Comparison" in worksheet.cell(2, headers["Combined details"]).value
    assert "Detail — Recorded ProcessTime1" not in headers
    assert "Detail — Adjusted Type I rate" not in headers
    workbook.close()
