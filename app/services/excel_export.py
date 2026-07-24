"""Signed, in-memory Excel export for one CryoCheck audit result."""

from __future__ import annotations

import hmac
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Final, Iterable

from itsdangerous import BadData, SignatureExpired, URLSafeTimedSerializer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.validation_engine import AuditException, AuditResult


_EXPORT_SALT: Final = "cryocheck-exception-export-v1"
_EXPORT_VERSION: Final = 1
_IDENTIFIER_PATTERN: Final = re.compile(r"^exception-[1-9][0-9]*$")
_FORMULA_PREFIXES: Final = ("=", "+", "-", "@")
_SOURCE_COLUMNS: Final[tuple[tuple[str, str], ...]] = (
    ("CSV source row", "source_row_number"),
    ("Rule ID", "rule_id"),
    ("Rule name", "rule_name"),
    ("Exception message", "exception_message"),
    ("Active settings profile", "active_settings_profile_name"),
    ("RecordID", "record_id"),
    ("ApplicationNumber", "application_number"),
    ("Gateway", "gateway_code"),
    ("AircraftType", "aircraft_type"),
    ("TailNumber", "tail_number"),
    ("ApplicationDate", "application_date"),
    ("StartTime", "start_time"),
    ("DateCreated", "date_created"),
    ("TruckNumber", "truck_number"),
    ("Operator", "operator"),
    ("Driver", "driver"),
)
_HEADER_FILL: Final = PatternFill("solid", fgColor="0B3558")
_HEADER_FONT: Final = Font(color="FFFFFF", bold=True)
_HEADER_ALIGNMENT: Final = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
_CELL_ALIGNMENT: Final = Alignment(vertical="top", wrap_text=True)


class ExportRequestError(ValueError):
    """A safe, user-facing export validation failure."""


@dataclass(frozen=True, slots=True)
class ExportRow:
    """One signed exception row eligible for export."""

    identifier: str
    source_row_number: int
    rule_id: str
    rule_name: str
    exception_message: str
    active_settings_profile_name: str
    record_id: str
    application_number: str
    gateway_code: str
    aircraft_type: str
    tail_number: str
    application_date: str
    start_time: str
    date_created: str
    truck_number: str
    operator: str
    driver: str
    details: tuple[tuple[str, str], ...]


@dataclass(frozen=True, slots=True)
class ExportSnapshot:
    """Validated signed export state from one Results page."""

    rows: tuple[ExportRow, ...]


@dataclass(frozen=True, slots=True)
class PreparedExport:
    """Signed form token and matching checkbox identifiers."""

    token: str
    identifiers: tuple[str, ...]


def prepare_export(
    audit: AuditResult,
    *,
    secret_key: str,
    context_id: str,
) -> PreparedExport:
    """Sign only exportable exceptions from one request-scoped audit."""
    if not context_id:
        raise RuntimeError("An export context is required.")
    rows = tuple(
        _export_row_from_exception(
            exception,
            identifier=f"exception-{index}",
            profile_name=audit.active_settings_profile_name,
        )
        for index, exception in enumerate(audit.exceptions, start=1)
    )
    payload = {
        "version": _EXPORT_VERSION,
        "context_id": context_id,
        "rows": [_row_to_payload(row) for row in rows],
    }
    token = _serializer(secret_key).dumps(payload)
    return PreparedExport(
        token=token,
        identifiers=tuple(row.identifier for row in rows),
    )


def load_export_snapshot(
    token: str,
    *,
    secret_key: str,
    max_age_seconds: int,
    expected_context_id: str,
) -> ExportSnapshot:
    """Verify token age, signature, and complete payload structure."""
    if not token:
        raise ExportRequestError(
            "The export request is missing its audit snapshot. "
            "Import the CSV again."
        )

    try:
        payload = _serializer(secret_key).loads(
            token,
            max_age=max_age_seconds,
        )
    except SignatureExpired as error:
        raise ExportRequestError(
            "This export request expired. Import the CSV again to create "
            "a fresh Results page."
        ) from error
    except BadData as error:
        raise ExportRequestError(
            "This export request is invalid. Import the CSV again."
        ) from error

    if (
        not isinstance(payload, dict)
        or not isinstance(payload.get("context_id"), str)
        or not expected_context_id
        or not hmac.compare_digest(
            payload["context_id"],
            expected_context_id,
        )
    ):
        raise ExportRequestError(
            "This export request is no longer the current audit result. "
            "Import the CSV again."
        )

    try:
        return _snapshot_from_payload(payload)
    except (KeyError, TypeError, ValueError) as error:
        raise ExportRequestError(
            "This export request is malformed. Import the CSV again."
        ) from error


def select_export_rows(
    snapshot: ExportSnapshot,
    *,
    scope: str,
    selected_identifiers: Iterable[str],
) -> tuple[ExportRow, ...]:
    """Validate identifiers and retain the original audit ordering."""
    submitted = tuple(selected_identifiers)
    known_identifiers = tuple(row.identifier for row in snapshot.rows)
    known_set = set(known_identifiers)

    if len(submitted) != len(set(submitted)):
        raise ExportRequestError(
            "The export request contains duplicate exception selections."
        )
    if any(identifier not in known_set for identifier in submitted):
        raise ExportRequestError(
            "The export request contains an exception that is not part of "
            "this audit result."
        )

    if scope == "all":
        selected_rows = snapshot.rows
    elif scope == "selected":
        if not submitted:
            raise ExportRequestError(
                "Select at least one exception before exporting."
            )
        selected_set = set(submitted)
        selected_rows = tuple(
            row for row in snapshot.rows if row.identifier in selected_set
        )
    else:
        raise ExportRequestError(
            "Choose Export Selected or Export Exceptions."
        )

    if not selected_rows:
        raise ExportRequestError("This audit result has no exceptions to export.")
    return selected_rows


def build_exception_workbook(
    rows: tuple[ExportRow, ...],
    *,
    now: datetime | None = None,
) -> tuple[BytesIO, str]:
    """Create a styled XLSX workbook without filesystem access."""
    detail_labels = _ordered_detail_labels(rows)
    headers = (
        *(label for label, _attribute in _SOURCE_COLUMNS),
        *(f"Detail — {label}" for label in detail_labels),
        "Combined details",
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Exceptions"
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT

    for export_row in rows:
        detail_values = dict(export_row.details)
        source_values = tuple(
            getattr(export_row, attribute)
            for _label, attribute in _SOURCE_COLUMNS
        )
        combined_details = "; ".join(
            f"{label}: {value}" for label, value in export_row.details
        )
        values = (
            *source_values,
            *(detail_values.get(label, "") for label in detail_labels),
            combined_details,
        )
        worksheet.append(
            tuple(
                value
                if isinstance(value, int)
                else _safe_excel_text(str(value))
                for value in values
            )
        )

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _CELL_ALIGNMENT

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 34
    _set_readable_widths(worksheet, headers)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    generated_at = now or datetime.now(timezone.utc)
    filename = (
        "CryoCheck_Exceptions_"
        f"{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return output, filename


def _serializer(secret_key: str) -> URLSafeTimedSerializer:
    if not secret_key:
        raise RuntimeError("A secret key is required for exception exports.")
    return URLSafeTimedSerializer(secret_key, salt=_EXPORT_SALT)


def _export_row_from_exception(
    exception: AuditException,
    *,
    identifier: str,
    profile_name: str,
) -> ExportRow:
    return ExportRow(
        identifier=identifier,
        source_row_number=exception.source_row_number,
        rule_id=exception.rule_id,
        rule_name=exception.rule_name,
        exception_message=exception.exception_message,
        active_settings_profile_name=profile_name,
        record_id=exception.record_id,
        application_number=exception.application_number,
        gateway_code=exception.gateway_code,
        aircraft_type=exception.aircraft_type,
        tail_number=exception.tail_number,
        application_date=exception.application_date,
        start_time=exception.start_time,
        date_created=exception.date_created,
        truck_number=exception.truck_number,
        operator=exception.operator,
        driver=exception.driver,
        details=tuple(
            (detail.label, detail.value) for detail in exception.details
        ),
    )


def _row_to_payload(row: ExportRow) -> dict[str, object]:
    return {
        "identifier": row.identifier,
        "source_row_number": row.source_row_number,
        "rule_id": row.rule_id,
        "rule_name": row.rule_name,
        "exception_message": row.exception_message,
        "active_settings_profile_name": row.active_settings_profile_name,
        "record_id": row.record_id,
        "application_number": row.application_number,
        "gateway_code": row.gateway_code,
        "aircraft_type": row.aircraft_type,
        "tail_number": row.tail_number,
        "application_date": row.application_date,
        "start_time": row.start_time,
        "date_created": row.date_created,
        "truck_number": row.truck_number,
        "operator": row.operator,
        "driver": row.driver,
        "details": [list(detail) for detail in row.details],
    }


def _snapshot_from_payload(payload: object) -> ExportSnapshot:
    if not isinstance(payload, dict) or payload.get("version") != _EXPORT_VERSION:
        raise ValueError("Unsupported export snapshot version.")
    row_payloads = payload["rows"]
    if not isinstance(row_payloads, list):
        raise TypeError("Export rows must be a list.")

    rows = tuple(_row_from_payload(row_payload) for row_payload in row_payloads)
    identifiers = tuple(row.identifier for row in rows)
    if len(identifiers) != len(set(identifiers)):
        raise ValueError("Export identifiers must be unique.")
    return ExportSnapshot(rows=rows)


def _row_from_payload(payload: object) -> ExportRow:
    if not isinstance(payload, dict):
        raise TypeError("Export row must be an object.")

    identifier = _required_string(payload, "identifier")
    if _IDENTIFIER_PATTERN.fullmatch(identifier) is None:
        raise ValueError("Invalid export identifier.")

    source_row_number = payload["source_row_number"]
    if type(source_row_number) is not int or source_row_number < 2:
        raise ValueError("Invalid CSV source row.")

    details_payload = payload["details"]
    if not isinstance(details_payload, list):
        raise TypeError("Export details must be a list.")
    details: list[tuple[str, str]] = []
    seen_detail_labels: set[str] = set()
    for detail in details_payload:
        if (
            not isinstance(detail, list)
            or len(detail) != 2
            or not all(isinstance(value, str) for value in detail)
            or not detail[0]
            or detail[0] in seen_detail_labels
        ):
            raise ValueError("Invalid export detail.")
        seen_detail_labels.add(detail[0])
        details.append((detail[0], detail[1]))

    return ExportRow(
        identifier=identifier,
        source_row_number=source_row_number,
        rule_id=_required_string(payload, "rule_id"),
        rule_name=_required_string(payload, "rule_name"),
        exception_message=_required_string(payload, "exception_message"),
        active_settings_profile_name=_required_string(
            payload,
            "active_settings_profile_name",
        ),
        record_id=_string(payload, "record_id"),
        application_number=_string(payload, "application_number"),
        gateway_code=_string(payload, "gateway_code"),
        aircraft_type=_string(payload, "aircraft_type"),
        tail_number=_string(payload, "tail_number"),
        application_date=_string(payload, "application_date"),
        start_time=_string(payload, "start_time"),
        date_created=_string(payload, "date_created"),
        truck_number=_string(payload, "truck_number"),
        operator=_string(payload, "operator"),
        driver=_string(payload, "driver"),
        details=tuple(details),
    )


def _required_string(payload: dict[object, object], key: str) -> str:
    value = payload[key]
    if not isinstance(value, str) or not value:
        raise ValueError(f"{key} must be a nonblank string.")
    return value


def _string(payload: dict[object, object], key: str) -> str:
    value = payload[key]
    if not isinstance(value, str):
        raise TypeError(f"{key} must be a string.")
    return value


def _ordered_detail_labels(rows: tuple[ExportRow, ...]) -> tuple[str, ...]:
    labels: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for label, _value in row.details:
            if label not in seen:
                seen.add(label)
                labels.append(label)
    return tuple(labels)


def _safe_excel_text(value: str) -> str:
    if value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _set_readable_widths(worksheet, headers: tuple[str, ...]) -> None:
    for column_index, header in enumerate(headers, start=1):
        values = (
            str(worksheet.cell(row=row_index, column=column_index).value or "")
            for row_index in range(1, worksheet.max_row + 1)
        )
        content_width = max(len(value) for value in values)
        if header == "Combined details":
            width = min(max(content_width + 2, 36), 72)
        elif header.startswith("Detail — "):
            width = min(max(content_width + 2, 18), 42)
        else:
            width = min(max(content_width + 2, 12), 30)
        worksheet.column_dimensions[get_column_letter(column_index)].width = (
            width
        )


__all__ = [
    "ExportRequestError",
    "ExportRow",
    "ExportSnapshot",
    "PreparedExport",
    "build_exception_workbook",
    "load_export_snapshot",
    "prepare_export",
    "select_export_rows",
]
